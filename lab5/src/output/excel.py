import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter


def save_to_excel(df, output_path):
    df.to_excel(output_path, index=False, header=False)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    ws.title = 'Optional Task Output'

    for r in range(1, 3):
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")

    thick_blue = Side(border_style="thick", color="0000FF")
    
    rows = ws.max_row
    cols = ws.max_column

    for r in range(4, rows + 1):
        for c in range(1, cols + 1):
            border_args = {}
            if r == 4:
                border_args["top"] = thick_blue
            if r == rows:
                border_args["bottom"] = thick_blue
            if c == 1:
                border_args["left"] = thick_blue
            if c == cols:
                border_args["right"] = thick_blue
            if c != 4 and c < 6:
                border_args["right"] = thick_blue
            
            ws.cell(row=r, column=c).border = Border(**border_args)

            if c < 6:
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")
            else:
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    for c in range(1, cols + 1):
        col_letter = get_column_letter(c)
        if c == 2:
            ws.column_dimensions[col_letter].width = 11
        elif c == 4:
            ws.column_dimensions[col_letter].width = 3
        elif c < 6:
            ws.column_dimensions[col_letter].width = 7
        else:
            ws.column_dimensions[col_letter].width = 3

    wb.save(output_path)
